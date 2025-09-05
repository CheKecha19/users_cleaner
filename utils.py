# utils.py
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
from pathlib import Path
from config import SHTAT_DIR, KONTUR_DIR, DIADOC_DIR, ONEC_DIR, MAX_FILE_AGE_DAYS, MAX_ROWS
from datetime import datetime, timedelta

def is_file_recent(file_path):
    """Проверяет, актуален ли файл (создан/изменен не более MAX_FILE_AGE_DAYS дней назад)"""
    if not file_path.exists():
        return False
    
    file_mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
    return (datetime.now() - file_mtime) <= timedelta(days=MAX_FILE_AGE_DAYS)

def find_latest_file(directory, pattern):
    """Находит самый новый файл в директории, соответствующий шаблону"""
    files = []
    for file in directory.glob(pattern):
        if is_file_recent(file):
            files.append(file)
    
    if not files:
        return None
    
    return max(files, key=os.path.getmtime)

def get_onec_file():
    """Находит файл 1С"""
    return find_latest_file(ONEC_DIR, "*.xlsx")

def get_kontur_file():
    """Находит файл Контура"""
    return find_latest_file(KONTUR_DIR, "*.xlsx")

def get_diadoc_file():
    """Находит файл Диадока"""
    return find_latest_file(DIADOC_DIR, "*.xlsx")

def get_shtat_file():
    """Находит файл штатного расписания"""
    return find_latest_file(SHTAT_DIR, "*.xlsx")

def replace_yo(text):
    """Замена ё на е"""
    if pd.isna(text):
        return text
    return str(text).replace('ё', 'е').replace('Ё', 'Е')

def normalize_name(full_name):
    """Нормализация ФИО (извлечение имени и фамилии без отчества)"""
    if pd.isna(full_name):
        return ""
    
    name = replace_yo(str(full_name))
    parts = re.split(r'\s+', name.strip())
    
    if len(parts) >= 2:
        return f"{parts[0]} {parts[1]}".upper()
    elif len(parts) == 1:
        return parts[0].upper()
    return ""

def highlight_duplicates(df, column, duplicate_names, color='red'):
    """Подсветка дубликатов в DataFrame"""
    if color == 'red':
        fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    else:  # yellow
        fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    
    for idx, row in df.iterrows():
        normalized_name = normalize_name(row[column])
        if normalized_name in duplicate_names:
            yield fill
        else:
            yield None

def save_with_formatting(df, filename, sheet_name, highlighting):
    """Сохранение DataFrame с форматированием"""
    df.to_excel(filename, sheet_name=sheet_name, index=False)
    
    # Применяем форматирование
    wb = load_workbook(filename)
    ws = wb[sheet_name]
    
    for idx, fill in enumerate(highlighting, 1):
        if fill:
            for row in range(2, len(df) + 2):  # Начинаем со второй строки (первая - заголовок)
                ws.cell(row=row, column=idx).fill = fill
    
    wb.save(filename)

def load_kontur_data():
    """Загрузка данных из Контура"""
    try:
        kontur_file = get_kontur_file()
        if not kontur_file or not is_file_recent(kontur_file):
            print("Актуальный файл Контура не найден")
            return pd.DataFrame(columns=['Контур_ФИО', 'Контур_Администратор', 'Контур_статус'])
        
        # Читаем данные
        df = pd.read_excel(kontur_file)
        
        # Переименовываем колонки
        df = df.rename(columns={
            'ФИО': 'Контур_ФИО',
            'Администратор': 'Контур_Администратор',
            'Дата блокировки': 'Контур_статус'
        })
        
        # Создаем копию для безопасного изменения
        result_df = df[['Контур_ФИО', 'Контур_Администратор', 'Контур_статус']].copy()
        
        # Преобразуем булевы значения в "да"/"нет" для Контур_Администратор
        if 'Контур_Администратор' in result_df.columns:
            # Преобразуем в строки и применяем логику
            admin_series = result_df['Контур_Администратор'].astype(str)
            admin_series = admin_series.apply(
                lambda x: 'да' if x.lower() in ['true', 'истина', '1', 'yes', 'да'] 
                else 'нет' if x.lower() in ['false', 'ложь', '0', 'no', 'нет'] 
                else x
            )
            result_df = result_df.assign(Контур_Администратор=admin_series)
        
        # Преобразуем даты блокировки в статусы для Контур_статус
        if 'Контур_статус' in result_df.columns:
            # Правильная логика: если в ячейке есть данные (не пустая и не NaN) - пользователь заблокирован
            # Если ячейка пустая или NaN - пользователь активен
            status_series = result_df['Контур_статус'].apply(
                lambda x: 'заблокирована' if pd.notna(x) and str(x).strip() != '' 
                else 'активна'
            )
            result_df = result_df.assign(Контур_статус=status_series)
        
        return result_df
        
    except Exception as e:
        print(f"Ошибка при загрузке данных Контура: {e}")
        return pd.DataFrame(columns=['Контур_ФИО', 'Контур_Администратор', 'Контур_статус'])
    
def load_diadoc_data():
    """Загрузка данных из Диадока"""
    try:
        diadoc_file = get_diadoc_file()
        if not diadoc_file or not is_file_recent(diadoc_file):
            print("Актуальный файл Диадока не найден")
            return pd.DataFrame(columns=['Диадок_ФИО', 'Диадок_Активен', 'Диадок_Администратор'])
        
        df = pd.read_excel(diadoc_file)
        # Переименовываем колонки для удобства
        df = df.rename(columns={
            'ФИО': 'Диадок_ФИО',
            'Активен': 'Диадок_Активен',
            'Администратор': 'Диадок_Администратор'
        })
        return df[['Диадок_ФИО', 'Диадок_Активен', 'Диадок_Администратор']]
    except Exception as e:
        print(f"Ошибка при загрузке данных Диадока: {e}")
        return pd.DataFrame(columns=['Диадок_ФИО', 'Диадок_Активен', 'Диадок_Администратор'])

def load_shtat_data():
    """Загрузка данных из штатного расписания"""
    try:
        shtat_file = get_shtat_file()
        if not shtat_file or not is_file_recent(shtat_file):
            print("Актуальный файл штатного расписания не найден")
            return pd.DataFrame(columns=['Штатное_ФИО'])
        
        df = pd.read_excel(shtat_file)
        # Переименовываем колонки для удобства
        df = df.rename(columns={'Ф.И.О.': 'Штатное_ФИО'})
        return df[['Штатное_ФИО']]
    except Exception as e:
        print(f"Ошибка при загрузке данных штатного расписания: {e}")
        return pd.DataFrame(columns=['Штатное_ФИО'])

def load_onec_data():
    """Загрузка данных из 1С"""
    try:
        onec_file = get_onec_file()
        if not onec_file or not is_file_recent(onec_file):
            print("Актуальный файл 1С не найден")
            return pd.DataFrame(columns=['1C_ФИО', '1C_Активен'])
        
        # Читаем файл, пропускаем первые 3 строки (заголовки)
        df = pd.read_excel(onec_file, skiprows=3)
        
        # Переименовываем колонки для удобства
        df = df.rename(columns={
            'Полное имя': '1C_ФИО',
            'Вход в приложение разрешен': '1C_Активен'
        })
        
        # Оставляем только нужные колонки и фильтруем пустые значения
        df = df[['1C_ФИО', '1C_Активен']].dropna(subset=['1C_ФИО'])
        
        # Преобразуем статус активности в понятный формат
        df_processed = df.copy()
        df_processed.loc[:, '1C_Активен'] = df_processed['1C_Активен'].apply(
            lambda x: 'Да' if pd.notna(x) and str(x).strip() != '' else 'Нет'
        )
        
        return df_processed
    except Exception as e:
        print(f"Ошибка при загрузке данных 1С: {e}")
        return pd.DataFrame(columns=['1C_ФИО', '1C_Активен'])

def create_comparison_sheet(ad_employees, shtat_employees, filename):
    """Создание листа сравнения AD и Штатного расписания"""
    if not shtat_employees:
        return 0
    
    # Находим сотрудников, которые есть в AD, но нет в штатном расписании
    ad_set = set(normalize_name(name) for name in ad_employees)
    shtat_set = set(normalize_name(name) for name in shtat_employees)
    
    missing_in_shtat = ad_set - shtat_set
    
    # Создаем DataFrame для результатов сравнения
    comparison_data = []
    for name in missing_in_shtat:
        # Находим оригинальное имя из AD
        original_name = next((n for n in ad_employees if normalize_name(n) == name), name)
        comparison_data.append({
            'ФИО_AD': original_name,
            'Статус': 'Активен в AD, но отсутствует в штатном расписании'
        })
    
    comparison_df = pd.DataFrame(comparison_data)
    
    # Сохраняем в файл
    with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
        comparison_df.to_excel(writer, sheet_name='сравнение AD и Штатки', index=False)
    
    return len(missing_in_shtat)

def find_duplicates(df1, df2, col1, col2):
    """Поиск дубликатов между двумя DataFrame"""
    names1 = set(df1[col1].apply(normalize_name).dropna())
    names2 = set(df2[col2].apply(normalize_name).dropna())
    
    return names1.intersection(names2)

def find_internal_duplicates(df, column):
    """Поиск дубликатов внутри одного столбца"""
    normalized_names = df[column].apply(normalize_name)
    value_counts = normalized_names.value_counts()
    return set(value_counts[value_counts > 1].index)

def find_users_to_remove(edo_df, staff_df, gph_df):
    """Поиск пользователей для удаления из ЭДО"""
    # Создаем объединенный набор всех valid names
    all_valid_names = set()
    
    if not staff_df.empty and 'AD_ФИО' in staff_df.columns:
        all_valid_names.update(staff_df['AD_ФИО'].apply(normalize_name).dropna())
    
    if not gph_df.empty and 'AD_ФИО' in gph_df.columns:
        all_valid_names.update(gph_df['AD_ФИО'].apply(normalize_name).dropna())
    
    users_to_remove = []
    
    for _, row in edo_df.iterrows():
        # Первый столбец - ФИО
        fio_column = edo_df.columns[0]
        if pd.isna(row[fio_column]):
            continue
            
        normalized_name = normalize_name(row[fio_column])
        
        # Проверяем условия для удаления (нет в AD и активен/не заблокирован)
        if normalized_name not in all_valid_names:
            # Для Контура проверяем статус
            if 'Контур_статус' in edo_df.columns and row['Контур_статус'] == 'активна':
                users_to_remove.append(row)
            # Для Диадока проверяем активность
            elif 'Диадок_Активен' in edo_df.columns and row['Диадок_Активен'] == 'Да':
                users_to_remove.append(row)
            # Для 1С проверяем активность
            elif '1C_Активен' in edo_df.columns and row['1C_Активен'] == 'Да':
                users_to_remove.append(row)
    
    return pd.DataFrame(users_to_remove)